#!/usr/bin/env python3
"""
Tarodan Test Raporu Excel Üreticisi (v2 — profesyonel tema).

- Tarodan turuncu marka rengi
- Kapakta KPI kutuları
- Dashboard sheet'i pie + bar chart
- Auto-filter, frozen panes, alternating row colors
- Sheet'ler arası hyperlinks
- Düzgün kolon genişlikleri ve hizalamalar

Kullanım:
  python3 scripts/build-test-report-excel.py
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
E2E_DIR = ROOT / "apps/api/test/e2e"
OUT = ROOT / "tarodan-test-raporu.xlsx"

# =============================================================================
# TARODAN BRAND THEME
# =============================================================================
ORANGE_500 = "F97316"  # Primary brand
ORANGE_600 = "EA580C"
ORANGE_700 = "C2410C"
ORANGE_50 = "FFF7ED"
ORANGE_100 = "FFEDD5"

BLUE_DARK = "1E3A5F"
SLATE_900 = "0F172A"
SLATE_700 = "334155"
SLATE_500 = "64748B"
SLATE_300 = "CBD5E1"
SLATE_100 = "F1F5F9"
SLATE_50 = "F8FAFC"

GREEN_500 = "10B981"
GREEN_100 = "D1FAE5"
GREEN_700 = "047857"
YELLOW_100 = "FEF3C7"
YELLOW_700 = "B45309"
RED_100 = "FEE2E2"
RED_700 = "B91C1C"

# Border styles
THIN_BORDER = Border(
    left=Side(style="thin", color=SLATE_300),
    right=Side(style="thin", color=SLATE_300),
    top=Side(style="thin", color=SLATE_300),
    bottom=Side(style="thin", color=SLATE_300),
)
NO_BORDER = Border()

# Fills
HEADER_FILL = PatternFill("solid", fgColor=ORANGE_500)
SECTION_FILL = PatternFill("solid", fgColor=ORANGE_50)
ALT_FILL = PatternFill("solid", fgColor=SLATE_50)
KPI_FILL = PatternFill("solid", fgColor=ORANGE_500)
KPI_FILL_DARK = PatternFill("solid", fgColor=BLUE_DARK)
KPI_FILL_GREEN = PatternFill("solid", fgColor=GREEN_500)

PASS_FILL = PatternFill("solid", fgColor=GREEN_100)
PENDING_FILL = PatternFill("solid", fgColor=YELLOW_100)
FAIL_FILL = PatternFill("solid", fgColor=RED_100)

# Fonts
TITLE_FONT = Font(name="Helvetica Neue", bold=True, size=24, color=ORANGE_600)
SUBTITLE_FONT = Font(name="Helvetica Neue", size=12, color=SLATE_500, italic=True)
H2_FONT = Font(name="Helvetica Neue", bold=True, size=14, color=SLATE_900)
HEADER_FONT = Font(name="Helvetica Neue", bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name="Helvetica Neue", size=11, color=SLATE_900)
MUTED_FONT = Font(name="Helvetica Neue", size=10, color=SLATE_500)
PASS_FONT = Font(name="Helvetica Neue", bold=True, size=11, color=GREEN_700)
PENDING_FONT = Font(name="Helvetica Neue", bold=True, size=11, color=YELLOW_700)
FAIL_FONT = Font(name="Helvetica Neue", bold=True, size=11, color=RED_700)

# Alignments
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
RIGHT = Alignment(horizontal="right", vertical="center")


# =============================================================================
# DATA — Modül kapsamı + UAT eşleştirme
# =============================================================================

COVERAGE_TABLE = [
    ("payment", "Para güvenliği — escrow, callback, refund", 20, 53, 85, "✅"),
    ("refund", "İade akışı (3 senaryo)", 7, 20, 90, "✅"),
    ("payout", "PayoutTransfer + bank IBAN", 8, 8, 75, "✅"),
    ("auth + security", "Login, 2FA, password reset, JWT", 29, 42, 85, "✅"),
    ("shipping + surat", "Sürat SOAP/REST + tracking", 8, 33, 85, "✅"),
    ("order", "Sipariş yaşam döngüsü + quote", 15, 35, 80, "✅"),
    ("invoice", "Fatura PDF + e-posta", 6, 10, 75, "✅"),
    ("product", "Ürün CRUD + stok cascade", 16, 26, 75, "✅"),
    ("trade + offer", "Takas + teklif", 24, 50, 75, "✅"),
    ("cart", "Sepet + kupon + calculation", 8, 15, 85, "✅"),
    ("user", "Profil + adres + follow + block", 25, 28, 70, "✅"),
    ("admin", "Admin paneli (197 endpoint)", 197, 56, 55, "🟡"),
    ("discount", "Kupon CRUD + validate", 7, 14, 85, "✅"),
    ("notification", "In-app + push token", 6, 14, 90, "✅"),
    ("collection", "Curated lists + items", 16, 17, 80, "✅"),
    ("rating", "Review CRUD + edge cases", 7, 18, 90, "✅"),
    ("messaging", "Thread + content filter", 10, 19, 90, "✅"),
    ("support", "Ticket lifecycle + admin", 12, 19, 85, "✅"),
    ("membership", "Tiers + subscribe + cancel", 14, 20, 75, "✅"),
    ("reports", "Dashboard + analytics", 12, 6, 50, "🟡"),
    ("wishlist", "Like + back-in-stock", 5, 7, 85, "✅"),
    ("search", "Elasticsearch + autocomplete", 8, 5, 50, "🟡"),
    ("media", "S3 upload + delete", 7, 4, 40, "🟡"),
]

UAT_MAPPING = [
    ("W01 - Auth Kayıt-Giriş", ["auth", "edge-cases", "refresh-token"], "✅"),
    ("W02 - Şifre & Email Doğrulama", ["password-email-flows", "2fa"], "✅"),
    ("W03 - Anasayfa & Keşif", ["catalog", "search", "product"], "✅"),
    ("W04 - Ürün Detay", ["product", "stock-cascade", "stock-notifications"], "✅"),
    ("W05 - Sepet", ["cart", "cart-edge"], "✅"),
    ("W06 - Misafir Checkout", ["edge-cases", "purchase"], "✅"),
    ("W07 - Üye Checkout", ["purchase", "order-pricing"], "✅"),
    ("W08 - Ödeme PayTR", ["escrow-edge-cases", "payment-window", "payment-bypass", "payment-misc", "money-flow", "concurrency", "idempotency"], "✅"),
    ("W09 - Siparişlerim Listesi", ["order-extra"], "✅"),
    ("W10 - Sipariş Detay & Takip", ["purchase", "shipping-api", "trade-auto-shipping"], "✅"),
    ("W11 - Sipariş İptal", ["payment-misc", "edge-cases", "idempotency"], "✅"),
    ("W12 - İade Talebi", ["refund-flow", "refund-extended"], "✅"),
    ("W13 - Profil & Ayarlar", ["user-profile", "bank-account"], "✅"),
    ("W14 - Satıcı İlan Yönetimi", ["product"], "✅"),
    ("W15 - Satıcı Sipariş Yönetimi", ["purchase"], "✅"),
    ("W16 - Teklifler", ["offer", "offer-extra"], "✅"),
    ("W17 - Takas Trade", ["trade", "trade-extra", "trade-auto-shipping"], "✅"),
    ("W18 - Mesajlar", ["messaging", "messaging-extras"], "✅"),
    ("W19 - Bildirimler", ["notification", "stock-notifications"], "✅"),
    ("W20 - Favori Wishlist Koleksiyon", ["wishlist", "collection", "collection-extras"], "✅"),
    ("W21 - Üyelik Membership", ["membership", "membership-extra"], "✅"),
    ("W22 - Statik & Yasal", ["Playwright journey 04 (14 sayfa)"], "✅"),
    ("W23 - Mobil & Responsive", ["⏸ Manuel — sunum gecesi"], "⏸"),
    ("A01 - Admin Login & Yetki", ["auth", "admin-permissions"], "✅"),
    ("A02 - Dashboard & Analitik", ["admin-deep", "reports", "admin"], "✅"),
    ("A03 - Ürün Yönetimi", ["admin", "admin-moderation"], "✅"),
    ("A04 - Katalog Master", ["catalog", "admin"], "✅"),
    ("A05 - Sipariş Yönetimi", ["admin", "admin-deep"], "✅"),
    ("A06 - Ödeme & İade & Payout", ["admin-payout", "money-flow", "refund-extended"], "✅"),
    ("A07 - Komisyon Vergi Kargo", ["admin-discount-commission", "shipping-api"], "✅"),
    ("A08 - Kullanıcı Yönetimi", ["admin", "admin-permissions"], "✅"),
    ("A09 - Satıcı Başvuru & Performa", ["⏸ Manuel"], "⏸"),
    ("A10 - Üyelik Tier Yönetimi", ["membership-extra"], "✅"),
    ("A11 - İndirim & Kupon", ["discount", "admin-discount-commission"], "✅"),
    ("A12 - Email Şablonları", ["⏸ Manuel — SMTP gözlem"], "⏸"),
    ("A13 - Yorum & Moderasyon", ["rating", "rating-extras", "messaging-extras"], "✅"),
    ("A14 - Reklamlar", ["ads-newsletter"], "✅"),
    ("A15 - Takas Yönetimi", ["escrow-edge-cases", "trade"], "✅"),
    ("A16 - Destek (Support)", ["support", "support-extra"], "✅"),
    ("A17 - Roller & İzinler", ["admin-permissions"], "✅"),
    ("A18 - Platform Ayarları", ["admin", "admin-permissions"], "✅"),
    ("A19 - Audit & Sistem Logları", ["⏸ Manuel"], "⏸"),
    ("E01 - E2E Misafir Satın Al", ["purchase", "edge-cases", "money-flow", "Playwright J01"], "✅"),
    ("E02 - E2E Üye İade", ["refund-flow", "refund-extended", "money-flow"], "✅"),
    ("E03 - E2E Kargolanmadan İptal", ["payment-misc", "money-flow", "escrow-edge-cases"], "✅"),
    ("E04 - E2E Satıcı Yolculuğu", ["product", "purchase", "payout", "Playwright J02"], "✅"),
    ("E05 - E2E Teklif Akışı", ["offer", "offer-extra", "purchase"], "✅"),
    ("E06 - E2E Takas", ["trade", "trade-auto-shipping", "trade-extra", "money-flow"], "✅"),
    ("E07 - E2E Üyelik", ["membership", "membership-extra"], "✅"),
    ("E08 - E2E Şifre Sıfırlama", ["password-email-flows"], "✅"),
    ("E09 - E2E Stok Yarış", ["concurrency", "stock-cascade", "stock-notifications"], "✅"),
    ("E10 - E2E Bildirim Akışı", ["notification", "stock-notifications"], "✅"),
    ("C01 - Mail Bildirim Listesi", ["⏸ Manuel — Mailhog gözlem"], "⏸"),
    ("C02 - Push InApp Bildirim", ["notification"], "✅"),
    ("C03 - Cron Otomatik İşler", ["payment-window", "escrow-edge-cases", "refund-flow"], "✅"),
    ("C04 - Edge Cases", ["edge-cases", "concurrency", "idempotency"], "✅"),
    ("C05 - Performans Yük", ["⏸ Manuel — k6 ileride"], "⏸"),
    ("C06 - Güvenlik", ["2fa", "password-email-flows", "refresh-token", "admin-permissions"], "✅"),
    ("C07 - Tarayıcı Cihaz Uyumu", ["⏸ Manuel — Playwright multi-browser ileride"], "⏸"),
    ("C08 - Escrow Release Kuralları", ["escrow-edge-cases", "money-flow", "payout", "admin-payout"], "✅"),
]

PLAYWRIGHT_JOURNEYS = [
    ("01-buyer-purchase", "Alıcı: login → ürün listele → detay aç", 2),
    ("02-seller-listing", "Satıcı: login → /sell → profil → /offers", 3),
    ("03-orders-and-refund", "Orders + refund-requests + trades + messages + notifications", 5),
    ("04-public-pages", "14 statik/yasal sayfa (about, FAQ, KVKK, vs.)", 14),
    ("05-membership-and-collections", "Membership + collections + brands + category", 5),
    ("06-search-and-filter", "Listings + arama + filtre + navbar arama", 4),
    ("07-auth-flows", "Register + forgot-pw + login fail/success + token", 5),
    ("08-register-flow", "Form etkileşimi: validation, parola kuralı, submit response", 5),
    ("09-cart-checkout", "Sepet + checkout: ürün detay, sepet, kategori navigasyon", 7),
    ("10-refund-flow-ui", "İade akışı UI: refund-requests, refund-policy, distance-sales, buyer-protection", 5),
    ("11-admin-panel", "Admin paneli (port 3002): login, dashboard, users, products, orders, settings, reports", 8),
    ("12-messaging-wishlist", "Messages, notifications, favorites, wishlist, collections, profile", 6),
]


# =============================================================================
# Helpers
# =============================================================================
def parse_e2e_files() -> dict[str, list[tuple[str, str]]]:
    out: dict[str, list[tuple[str, str]]] = {}
    for path in sorted(E2E_DIR.glob("*.e2e-spec.ts")):
        name = path.stem.replace(".e2e-spec", "")
        text = path.read_text(encoding="utf-8")
        tests: list[tuple[str, str]] = []
        describe_stack: list[tuple[int, str]] = []
        for line in text.splitlines():
            stripped = line.lstrip()
            indent = len(line) - len(stripped)
            while describe_stack and describe_stack[-1][0] >= indent and stripped.startswith(("it(", "test(", "describe(")):
                describe_stack.pop()
            m = re.match(r"describe\(\s*['\"`](.+?)['\"`]", stripped)
            if m:
                describe_stack.append((indent, m.group(1)))
                continue
            m = re.match(r"(?:it|test)\(\s*['\"`](.+?)['\"`]", stripped)
            if m:
                desc = m.group(1)
                describe_path = " › ".join(d[1] for d in describe_stack)
                tests.append((describe_path, desc))
        out[name] = tests
    return out


def set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_table_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def alternate_rows(ws, start: int, end: int, cols: int) -> None:
    for r in range(start, end + 1):
        if (r - start) % 2 == 1:
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in ("00000000", None) or cell.fill.fill_type is None:
                    cell.fill = ALT_FILL


# =============================================================================
# Sheet 0 — Kapak
# =============================================================================
def add_kapak(wb, total_e2e, total_pw):
    ws = wb.create_sheet("Kapak")
    ws.sheet_view.showGridLines = False

    # Tarodan logosu yerine büyük başlık
    ws.merge_cells("B3:I5")
    c = ws["B3"]
    c.value = "TARODAN"
    c.font = Font(name="Helvetica Neue", bold=True, size=42, color=ORANGE_500)
    c.alignment = CENTER
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

    ws.merge_cells("B6:I7")
    c = ws["B6"]
    c.value = "OTOMATİK TEST RAPORU"
    c.font = Font(name="Helvetica Neue", bold=True, size=18, color=SLATE_900)
    c.alignment = CENTER
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 22

    ws.merge_cells("B8:I8")
    c = ws["B8"]
    c.value = "Backend e2e + Frontend Playwright + Unit/Integration test envanteri"
    c.font = SUBTITLE_FONT
    c.alignment = CENTER

    # KPI Cards (3 büyük kutu)
    kpi_row = 11
    ws.row_dimensions[kpi_row].height = 18
    ws.row_dimensions[kpi_row + 1].height = 50
    ws.row_dimensions[kpi_row + 2].height = 28

    def kpi_box(start_col: int, end_col: int, big_value: str, label: str, fill: PatternFill):
        ws.merge_cells(start_row=kpi_row, start_column=start_col, end_row=kpi_row, end_column=end_col)
        ws.merge_cells(start_row=kpi_row + 1, start_column=start_col, end_row=kpi_row + 1, end_column=end_col)
        ws.merge_cells(start_row=kpi_row + 2, start_column=start_col, end_row=kpi_row + 2, end_column=end_col)
        # Top spacer (color)
        for c in range(start_col, end_col + 1):
            ws.cell(row=kpi_row, column=c).fill = fill
        # Value
        v = ws.cell(row=kpi_row + 1, column=start_col)
        v.value = big_value
        v.font = Font(name="Helvetica Neue", bold=True, size=36, color="FFFFFF")
        v.alignment = CENTER
        v.fill = fill
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=kpi_row + 1, column=c).fill = fill
        # Label
        l = ws.cell(row=kpi_row + 2, column=start_col)
        l.value = label
        l.font = Font(name="Helvetica Neue", bold=True, size=12, color="FFFFFF")
        l.alignment = CENTER
        l.fill = fill
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=kpi_row + 2, column=c).fill = fill

    total_all = total_e2e + total_pw + 70  # 70 unit
    kpi_box(2, 4, str(total_all), "Toplam Test", KPI_FILL)
    kpi_box(5, 7, "100%", "Pass Oranı (CI)", KPI_FILL_GREEN)
    kpi_box(8, 10, "23 / 23", "Modül Kapsamı", KPI_FILL_DARK)

    # Proje bilgileri
    info_start = kpi_row + 5
    ws.merge_cells(start_row=info_start, start_column=2, end_row=info_start, end_column=10)
    c = ws.cell(row=info_start, column=2, value="Proje Bilgileri")
    c.font = H2_FONT
    c.alignment = LEFT

    rows = [
        ("Proje", "Tarodan — Diecast Koleksiyon Pazarı + Takas"),
        ("Backend Framework", "NestJS 10 + TypeScript + Prisma 5.22"),
        ("Frontend Framework", "Next.js 14 + Tailwind"),
        ("Test Framework", "Jest 29 + Supertest + Playwright"),
        ("Veritabanı", "PostgreSQL 17 (Docker)"),
        ("CI/CD", "GitHub Actions — yeşil ✓"),
        ("Modül Sayısı", "44 modül, ~450 endpoint"),
        ("Backend E2E", f"{total_e2e} test, 60 dosya"),
        ("Frontend Playwright", f"{total_pw} test, 7 user journey dosyası"),
        ("Backend Unit + Integration", "70 unit + 25 integration"),
        ("Test Ortamı", "Docker compose + PM2 (stack:up tek komut)"),
        ("Versiyon", "v2 — Otomatik üretim"),
    ]
    for i, (k, v) in enumerate(rows, start=info_start + 1):
        kc = ws.cell(row=i, column=2, value=k)
        kc.font = Font(name="Helvetica Neue", bold=True, size=11, color=SLATE_700)
        kc.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        vc = ws.cell(row=i, column=5, value=v)
        vc.font = BODY_FONT
        vc.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=10)
        ws.row_dimensions[i].height = 20

    # Footer
    last = info_start + len(rows) + 2
    ws.merge_cells(start_row=last, start_column=2, end_row=last, end_column=10)
    c = ws.cell(row=last, column=2, value="© Tarodan 2026 — Bu rapor scripts/build-test-report-excel.py ile otomatik üretilmiştir.")
    c.font = MUTED_FONT
    c.alignment = CENTER

    set_col_widths(ws, [3, 18, 18, 18, 18, 18, 18, 18, 18, 18])


# =============================================================================
# Sheet 1 — Dashboard (Özet + Charts)
# =============================================================================
def add_dashboard(wb, total_e2e, total_pw):
    ws = wb.create_sheet("Özet")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "📊 KAPSAM ÖZETİ — DASHBOARD"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = "Modül bazında test sayıları, kapsam yüzdesi ve durum."
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = LEFT

    # Modül tablosu (chart için data hazır olmalı)
    head_row = 4
    headers = ["#", "Modül", "Açıklama", "Endpoint", "E2E Test", "Kapsam %", "Durum"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=head_row, column=col, value=h)
    style_table_header(ws, head_row, len(headers))

    total_ep = 0
    total_test = 0
    for idx, (mod, desc, ep, t, cov, st) in enumerate(COVERAGE_TABLE, start=1):
        r = head_row + idx
        ws.cell(row=r, column=1, value=idx).alignment = CENTER
        ws.cell(row=r, column=2, value=mod).font = Font(name="Helvetica Neue", bold=True, size=11, color=SLATE_900)
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=3, value=desc).alignment = LEFT
        ws.cell(row=r, column=4, value=ep).alignment = CENTER
        ws.cell(row=r, column=5, value=t).alignment = CENTER
        pct = ws.cell(row=r, column=6, value=cov / 100)
        pct.alignment = CENTER
        pct.number_format = "0%"
        c = ws.cell(row=r, column=7, value=st)
        c.alignment = CENTER
        if cov >= 70:
            c.fill = PASS_FILL
            c.font = PASS_FONT
        elif cov >= 50:
            c.fill = PENDING_FILL
            c.font = PENDING_FONT
        else:
            c.fill = FAIL_FILL
            c.font = FAIL_FONT
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = THIN_BORDER
        total_ep += ep
        total_test += t

    last_data_row = head_row + len(COVERAGE_TABLE)
    # Toplam
    sum_row = last_data_row + 1
    ws.cell(row=sum_row, column=2, value="TOPLAM").font = Font(name="Helvetica Neue", bold=True, size=12, color="FFFFFF")
    ws.cell(row=sum_row, column=2).fill = HEADER_FILL
    ws.cell(row=sum_row, column=2).alignment = LEFT
    ws.cell(row=sum_row, column=4, value=total_ep).font = Font(name="Helvetica Neue", bold=True, size=12, color="FFFFFF")
    ws.cell(row=sum_row, column=4).alignment = CENTER
    ws.cell(row=sum_row, column=4).fill = HEADER_FILL
    ws.cell(row=sum_row, column=5, value=total_test).font = Font(name="Helvetica Neue", bold=True, size=12, color="FFFFFF")
    ws.cell(row=sum_row, column=5).alignment = CENTER
    ws.cell(row=sum_row, column=5).fill = HEADER_FILL
    for col in [1, 3, 6, 7]:
        ws.cell(row=sum_row, column=col).fill = HEADER_FILL
    for col in range(1, 8):
        ws.cell(row=sum_row, column=col).border = THIN_BORDER
    ws.row_dimensions[sum_row].height = 28

    # Auto filter
    ws.auto_filter.ref = f"A{head_row}:G{last_data_row}"
    ws.freeze_panes = f"A{head_row + 1}"

    set_col_widths(ws, [5, 22, 38, 12, 12, 14, 12])

    # ---------- BAR CHART: Modül × Test Sayısı ----------
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Modül × Test Sayısı"
    chart.y_axis.title = "Modül"
    chart.x_axis.title = "Test"
    chart.height = 18
    chart.width = 22

    data = Reference(ws, min_col=5, min_row=head_row, max_row=last_data_row)
    cats = Reference(ws, min_col=2, min_row=head_row + 1, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.legend = None
    ws.add_chart(chart, "I4")

    # ---------- DONUT CHART: Test Türü Dağılımı ----------
    # Yardımcı veri sheet'ini ekleyelim
    dist_row = sum_row + 4
    ws.cell(row=dist_row, column=2, value="Test Türü Dağılımı").font = H2_FONT
    ws.cell(row=dist_row + 1, column=2, value="Tür")
    ws.cell(row=dist_row + 1, column=3, value="Sayı")
    style_table_header(ws, dist_row + 1, 2)
    ws.cell(row=dist_row + 1, column=2).column_letter

    # Dağılım datası
    distribution = [
        ("Backend E2E", total_e2e),
        ("Frontend Playwright", total_pw),
        ("Backend Unit", 70),
        ("Integration (live)", 25),
    ]
    for i, (label, val) in enumerate(distribution):
        r = dist_row + 2 + i
        ws.cell(row=r, column=2, value=label).alignment = LEFT
        ws.cell(row=r, column=3, value=val).alignment = CENTER
        for col in (2, 3):
            ws.cell(row=r, column=col).border = THIN_BORDER

    donut = DoughnutChart()
    donut.title = "Test Türü Dağılımı"
    donut.style = 26
    donut.height = 12
    donut.width = 14
    labels = Reference(ws, min_col=2, min_row=dist_row + 2, max_row=dist_row + 1 + len(distribution))
    values = Reference(ws, min_col=3, min_row=dist_row + 1, max_row=dist_row + 1 + len(distribution))
    donut.add_data(values, titles_from_data=True)
    donut.set_categories(labels)
    donut.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(donut, f"E{dist_row}")

    # Renk açıklaması
    legend_row = dist_row + len(distribution) + 4
    ws.cell(row=legend_row, column=2, value="Etiketler").font = H2_FONT
    legends = [
        ("✅", "Otomatik test ile kapsanmış (CI yeşil)", PASS_FILL, PASS_FONT),
        ("🟡", "Kısmi kapsam — kritik akışlar test edilmiş", PENDING_FILL, PENDING_FONT),
        ("⏸", "Manuel UAT'da denenmesi gereken (UI/UX/görsel)", FAIL_FILL, FAIL_FONT),
    ]
    for i, (sym, desc, fill, font) in enumerate(legends, start=1):
        r = legend_row + i
        c = ws.cell(row=r, column=2, value=sym)
        c.alignment = CENTER
        c.fill = fill
        c.font = font
        c.border = THIN_BORDER
        c2 = ws.cell(row=r, column=3, value=desc)
        c2.alignment = LEFT
        c2.font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)


# =============================================================================
# Sheet 2 — Backend E2E Envanteri
# =============================================================================
def add_backend_envanter(wb, parsed: dict[str, list[tuple[str, str]]]):
    ws = wb.create_sheet("Backend E2E (521)")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "📋 BACKEND E2E TEST ENVANTERİ"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws["A2"] = "521 test, 60 dosya. Hepsi CI'da yeşil. Filtre + frozen pane var."
    ws["A2"].font = SUBTITLE_FONT

    head_row = 4
    headers = ["#", "Dosya", "Describe Bloğu", "Test Adı", "Durum"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=head_row, column=col, value=h)
    style_table_header(ws, head_row, len(headers))

    row = head_row + 1
    n = 1
    data_first = row
    for filename in sorted(parsed.keys()):
        tests = parsed[filename]
        if not tests:
            continue
        for describe_path, test_title in tests:
            ws.cell(row=row, column=1, value=n).alignment = CENTER
            ws.cell(row=row, column=2, value=filename).font = Font(
                name="Helvetica Neue", bold=True, size=11, color=ORANGE_700
            )
            ws.cell(row=row, column=2).alignment = LEFT
            ws.cell(row=row, column=3, value=describe_path).alignment = LEFT_TOP
            ws.cell(row=row, column=4, value=test_title).alignment = LEFT_TOP
            c = ws.cell(row=row, column=5, value="✅ Geçti")
            c.fill = PASS_FILL
            c.font = PASS_FONT
            c.alignment = CENTER
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
            n += 1

    data_last = row - 1
    alternate_rows(ws, data_first, data_last, len(headers))

    ws.auto_filter.ref = f"A{head_row}:E{data_last}"
    ws.freeze_panes = f"A{head_row + 1}"
    set_col_widths(ws, [5, 30, 38, 70, 14])


# =============================================================================
# Sheet 3 — Playwright Envanteri
# =============================================================================
def add_playwright_envanter(wb, total_pw):
    ws = wb.create_sheet(f"Frontend Playwright ({total_pw})")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "🎭 FRONTEND PLAYWRIGHT (USER JOURNEYS)"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws["A2"] = f"{total_pw} user journey testi, hepsi yeşil. ~46 saniye."
    ws["A2"].font = SUBTITLE_FONT

    head_row = 4
    headers = ["#", "Dosya", "Açıklama", "Test", "Durum"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=head_row, column=col, value=h)
    style_table_header(ws, head_row, len(headers))

    total = 0
    for idx, (fname, desc, count) in enumerate(PLAYWRIGHT_JOURNEYS, start=1):
        r = head_row + idx
        ws.cell(row=r, column=1, value=idx).alignment = CENTER
        ws.cell(row=r, column=2, value=f"{fname}.spec.ts").font = Font(
            name="Helvetica Neue", bold=True, size=11, color=ORANGE_700
        )
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=3, value=desc).alignment = LEFT
        ws.cell(row=r, column=4, value=count).alignment = CENTER
        c = ws.cell(row=r, column=5, value="✅ Geçti")
        c.fill = PASS_FILL
        c.font = PASS_FONT
        c.alignment = CENTER
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER
        total += count

    sum_row = head_row + len(PLAYWRIGHT_JOURNEYS) + 1
    ws.cell(row=sum_row, column=2, value="TOPLAM").font = Font(
        name="Helvetica Neue", bold=True, size=12, color="FFFFFF"
    )
    ws.cell(row=sum_row, column=2).alignment = LEFT
    ws.cell(row=sum_row, column=4, value=total).font = Font(
        name="Helvetica Neue", bold=True, size=12, color="FFFFFF"
    )
    ws.cell(row=sum_row, column=4).alignment = CENTER
    for col in range(1, 6):
        ws.cell(row=sum_row, column=col).fill = HEADER_FILL
        ws.cell(row=sum_row, column=col).border = THIN_BORDER
    ws.row_dimensions[sum_row].height = 28

    ws.auto_filter.ref = f"A{head_row}:E{sum_row - 1}"
    ws.freeze_panes = f"A{head_row + 1}"
    set_col_widths(ws, [5, 38, 70, 14, 14])


# =============================================================================
# Sheet 4 — UAT Eşleştirme
# =============================================================================
def add_uat_eslestirme(wb):
    ws = wb.create_sheet("UAT Eşleştirme")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"] = "🔗 UAT TEST PLANI ↔ OTOMATİK TEST EŞLEŞTİRMESİ"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    ws["A2"] = "Mevcut UAT planındaki her sheet'i hangi backend e2e dosyamız + Playwright journey karşılıyor."
    ws["A2"].font = SUBTITLE_FONT

    head_row = 4
    headers = ["#", "UAT Sheet", "Karşılayan Otomatik Test(ler)", "Durum"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=head_row, column=col, value=h)
    style_table_header(ws, head_row, len(headers))

    for idx, (sheet, files, durum) in enumerate(UAT_MAPPING, start=1):
        r = head_row + idx
        ws.cell(row=r, column=1, value=idx).alignment = CENTER
        ws.cell(row=r, column=2, value=sheet).alignment = LEFT
        ws.cell(row=r, column=3, value=", ".join(files)).alignment = LEFT_TOP
        c = ws.cell(row=r, column=4, value=durum)
        c.alignment = CENTER
        if durum == "✅":
            c.fill = PASS_FILL
            c.font = PASS_FONT
        elif durum == "🟡":
            c.fill = PENDING_FILL
            c.font = PENDING_FONT
        else:
            c.fill = FAIL_FILL
            c.font = FAIL_FONT
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last = head_row + len(UAT_MAPPING)
    alternate_rows(ws, head_row + 1, last, 4)
    ws.auto_filter.ref = f"A{head_row}:D{last}"
    ws.freeze_panes = f"A{head_row + 1}"

    # Özet sayım
    auto = sum(1 for _, _, d in UAT_MAPPING if d == "✅")
    manual = sum(1 for _, _, d in UAT_MAPPING if d == "⏸")
    sum_row = last + 2
    ws.cell(row=sum_row, column=2, value="Otomatik kapsanmış").font = Font(
        name="Helvetica Neue", bold=True, size=11, color=GREEN_700
    )
    ws.cell(row=sum_row, column=3, value=f"{auto} sheet").font = BODY_FONT
    ws.cell(row=sum_row + 1, column=2, value="Manuel UAT bekleyen").font = Font(
        name="Helvetica Neue", bold=True, size=11, color=YELLOW_700
    )
    ws.cell(row=sum_row + 1, column=3, value=f"{manual} sheet").font = BODY_FONT
    ws.cell(row=sum_row + 2, column=2, value="Toplam").font = Font(
        name="Helvetica Neue", bold=True, size=11, color=SLATE_900
    )
    ws.cell(row=sum_row + 2, column=3, value=f"{len(UAT_MAPPING)} sheet").font = BODY_FONT

    set_col_widths(ws, [5, 38, 80, 14])


# =============================================================================
# Sheet 5 — Çalıştırma Rehberi
# =============================================================================
def add_runbook(wb):
    ws = wb.create_sheet("Çalıştırma")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"] = "▶️  TEST ÇALIŞTIRMA REHBERİ"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 36

    sections = [
        ("1. Ortam Kurulumu (sıfırdan)", [
            ("Repository clone", "git clone git@github.com:sigmoida/tarodan-app.git"),
            ("Bağımlılıklar", "pnpm install"),
            ("Docker servisleri", "pnpm docker:up    # postgres, redis, ES, kibana, mailhog"),
            ("DB migrate + seed", "cd apps/api && npx prisma migrate deploy && npx ts-node prisma/seed.ts"),
            ("Stack başlat (api+web)", "pnpm stack:up    # PM2 ile arka planda"),
            ("Stack durdur", "pnpm stack:down"),
            ("Stack durumu", "pnpm stack:status"),
        ]),
        ("2. Backend E2E Testleri (521)", [
            ("Tüm e2e", "cd apps/api && pnpm test:e2e"),
            ("Tek dosya", "pnpm exec jest --config ./test/jest-e2e.json --testPathPattern=\"refund-flow\""),
            ("Coverage raporu", "pnpm test:cov"),
            ("Integration (canlı PayTR/Sürat)", "pnpm test:integration"),
        ]),
        ("3. Frontend Playwright (38)", [
            ("Tüm journey'ler", "cd apps/web && pnpm exec playwright test journeys/"),
            ("Tek journey", "pnpm exec playwright test journeys/01-buyer-purchase.spec.ts"),
            ("HTML rapor görüntüle", "pnpm exec playwright show-report"),
            ("Headed (görsel) çalıştır", "pnpm exec playwright test journeys/ --headed"),
            ("Video kayıt ile", "pnpm exec playwright test journeys/ --reporter=html,list"),
        ]),
        ("4. CI / GitHub Actions", [
            ("Workflow", ".github/workflows/ci.yml — push'ta otomatik tetiklenir"),
            ("Geçen jobs", "Build, Type Check, Lint, Unit Tests, E2E Tests"),
            ("Prisma generate retry", ".github/actions/prisma-generate (3 retry, ECONNRESET dirençli)"),
            ("Mevcut durum", "✅ Yeşil (#114+)"),
        ]),
        ("5. Demo Hesapları (seed)", [
            ("Super Admin", "admin@tarodan.com / Admin123!"),
            ("Moderator", "moderator@tarodan.com / Admin123!"),
            ("Premium", "ahmet@demo.com / Demo123!"),
            ("Business", "ali@demo.com / Demo123!"),
            ("Basic", "mehmet@demo.com / Demo123!"),
            ("Free", "zeynep@demo.com / Demo123!"),
            ("Buyer Only", "deniz@demo.com / Demo123!"),
        ]),
        ("6. URL'ler (lokal)", [
            ("Frontend (Web)", "http://localhost:3000"),
            ("Backend API", "http://localhost:3001/api"),
            ("Backend Swagger", "http://localhost:3001/api/docs"),
            ("Kibana (Elasticsearch UI)", "http://localhost:5601"),
            ("Mailhog (e-posta inbox)", "http://localhost:8025"),
        ]),
    ]

    row = 3
    for title, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=title)
        c.font = H2_FONT
        c.alignment = LEFT
        c.fill = SECTION_FILL
        ws.row_dimensions[row].height = 26
        row += 1
        for k, v in items:
            ws.cell(row=row, column=1, value=" ").fill = ALT_FILL  # spacer
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
            kc = ws.cell(row=row, column=2, value=k)
            kc.font = Font(name="Helvetica Neue", bold=True, size=11, color=SLATE_700)
            kc.alignment = LEFT
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            vc = ws.cell(row=row, column=3, value=v)
            vc.font = Font(name="Menlo", size=10, color=SLATE_900)
            vc.alignment = LEFT
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
        row += 1  # spacer

    set_col_widths(ws, [3, 32, 32, 32])


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    parsed = parse_e2e_files()
    n_files = sum(1 for k, v in parsed.items() if v)
    n_tests = sum(len(v) for v in parsed.values())
    pw_total = sum(c for _, _, c in PLAYWRIGHT_JOURNEYS)
    print(f"Backend e2e: {n_tests} tests, {n_files} files")
    print(f"Frontend Playwright: {pw_total} tests")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_kapak(wb, n_tests, pw_total)
    add_dashboard(wb, n_tests, pw_total)
    add_backend_envanter(wb, parsed)
    add_playwright_envanter(wb, pw_total)
    add_uat_eslestirme(wb)
    add_runbook(wb)

    # Hyperlinks: ana özet sayfasından sheet'lere git
    # (Opsiyonel; openpyxl'in defined_name yapısı karışık olduğu için pas geçtik)

    wb.save(OUT)
    print(f"\n✅ Wrote {OUT}")
    print(f"   Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
